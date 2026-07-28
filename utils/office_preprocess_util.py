"""Structure-aware preprocessing for Excel and Word documents."""

from __future__ import annotations

import json
import re
from datetime import date, datetime, time
from pathlib import Path
from statistics import median
from typing import Any, Iterable, Iterator, Sequence
from uuid import uuid4

import pandas as pd
import xlrd
from docx import Document
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


FIXED_DELIMITER = "<FIXED_DELIMITER>"
DEFAULT_EXCLUDED_SHEETS = ("質問", "QA", "Q&A", "Answer", "Answers")
DEFAULT_MAX_BLOCK_CHARS = 3000
PROCEDURE_HEADER_HINTS = (
    "作業項目",
    "作業内容",
    "コマンド",
    "確認ポイント",
    "確認条件",
    "備考",
)


def _normalize_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (datetime, date, time, pd.Timestamp)):
        return value.isoformat()
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _is_empty(value: Any) -> bool:
    value = _normalize_value(value)
    return value is None or (isinstance(value, str) and not value.strip())


def _parse_excluded_sheets(excluded_sheets: str | Sequence[str] | None) -> set[str]:
    if excluded_sheets is None:
        excluded_sheets = DEFAULT_EXCLUDED_SHEETS
    if isinstance(excluded_sheets, str):
        excluded_sheets = re.split(r"[,，\n]", excluded_sheets)
    return {str(name).strip().casefold() for name in excluded_sheets if str(name).strip()}


def _json_text(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def _split_long_text(text: str, budget: int) -> list[str]:
    if len(text) <= budget:
        return [text]
    parts: list[str] = []
    current = ""
    for line in text.splitlines() or [text]:
        candidate = f"{current}\n{line}".strip() if current else line
        if len(candidate) <= budget:
            current = candidate
            continue
        if current:
            parts.append(current)
        if len(line) > budget:
            # Preserve an indivisible paragraph/command rather than cutting it
            # in the middle. Such a part may exceed the requested soft limit.
            parts.append(line)
            current = ""
        else:
            current = line
    if current:
        parts.append(current)
    return parts


def split_evidence_blocks(
        blocks: Iterable[dict[str, Any]],
        max_block_chars: int = DEFAULT_MAX_BLOCK_CHARS,
) -> list[dict[str, Any]]:
    """Keep serialized evidence blocks bounded while repeating provenance."""
    max_block_chars = max(512, int(max_block_chars or DEFAULT_MAX_BLOCK_CHARS))
    result: list[dict[str, Any]] = []
    for block in blocks:
        if len(_json_text(block)) <= max_block_chars:
            result.append(block)
            continue
        if block.get("block_type") == "word_table_row":
            result.append(block)
            continue

        base = {key: value for key, value in block.items() if key not in {"content", "fields"}}
        if "content" in block:
            budget = max(128, max_block_chars - len(_json_text(base)) - 120)
            payloads = _split_long_text(str(block["content"]), budget)
            parts = [{**base, "content": payload} for payload in payloads]
        else:
            lines = [f"{key}: {value}" for key, value in block.get("fields", {}).items()]
            budget = max(128, max_block_chars - len(_json_text(base)) - 120)
            payloads = _split_long_text("\n".join(lines), budget)
            parts = [{**base, "content": payload} for payload in payloads]

        for index, part in enumerate(parts, start=1):
            part["part_index"] = index
            part["part_count"] = len(parts)
            result.append(part)
    return result


def create_preprocessed_output_path(source_path: str | Path) -> Path:
    source = Path(source_path)
    output_dir = Path("output") / "preprocessed" / f"{source.stem}-{uuid4().hex[:8]}"
    output_dir.mkdir(parents=True, exist_ok=False)
    return output_dir / f"{source.name}.txt"


def write_evidence_document(blocks: Sequence[dict[str, Any]], output_path: str | Path) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8") as stream:
        for block in blocks:
            stream.write(f"{_json_text(block)} {FIXED_DELIMITER}\n")
    return output.resolve()


def _merged_value_map(worksheet) -> dict[tuple[int, int], tuple[int, int]]:
    merged: dict[tuple[int, int], tuple[int, int]] = {}
    for cell_range in worksheet.merged_cells.ranges:
        anchor = (cell_range.min_row, cell_range.min_col)
        for row in range(cell_range.min_row, cell_range.max_row + 1):
            for column in range(cell_range.min_col, cell_range.max_col + 1):
                merged[(row, column)] = anchor
    return merged


def _xlsx_rows(formula_sheet, value_sheet) -> list[list[Any]]:
    merged = _merged_value_map(formula_sheet)
    max_row = formula_sheet.max_row
    max_column = formula_sheet.max_column
    formula_rows = list(formula_sheet.iter_rows(
        min_row=1,
        max_row=max_row,
        max_col=max_column,
        values_only=True,
    ))
    cached_rows = list(value_sheet.iter_rows(
        min_row=1,
        max_row=max_row,
        max_col=max_column,
        values_only=True,
    ))
    rows: list[list[Any]] = []
    for row_index in range(1, max_row + 1):
        values: list[Any] = []
        for column_index in range(1, max_column + 1):
            source_row, source_column = merged.get(
                (row_index, column_index),
                (row_index, column_index),
            )
            # Horizontal merged labels belong to this logical row and may be
            # repeated safely. A vertically merged value must not leak into a
            # later table row or create an extra procedure-step boundary.
            if source_row != row_index:
                values.append(None)
                continue
            formula = formula_rows[source_row - 1][source_column - 1]
            cached = cached_rows[source_row - 1][source_column - 1]
            values.append(_normalize_value(cached if cached is not None else formula))
        rows.append(values)
    return rows


def _dataframe_rows(frame: pd.DataFrame) -> list[list[Any]]:
    return [[_normalize_value(value) for value in row] for row in frame.itertuples(index=False, name=None)]


def _xls_rows(sheet, datemode: int) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for row_index in range(sheet.nrows):
        values: list[Any] = []
        for column_index in range(sheet.ncols):
            cell = sheet.cell(row_index, column_index)
            value = cell.value
            if cell.ctype == xlrd.XL_CELL_DATE:
                value = xlrd.xldate_as_datetime(value, datemode)
            elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                value = bool(value)
            elif cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
                value = None
            values.append(_normalize_value(value))
        rows.append(values)
    return rows


def _nonempty_rows_after(rows: Sequence[Sequence[Any]], start: int, limit: int = 50):
    selected = []
    for index in range(start, len(rows)):
        if any(not _is_empty(value) for value in rows[index]):
            selected.append((index, rows[index]))
            if len(selected) >= limit:
                break
    return selected


def _find_header_row(rows: Sequence[Sequence[Any]]) -> int | None:
    best_index = None
    best_score = float("-inf")
    for index, row in enumerate(rows[:30]):
        populated = [column for column, value in enumerate(row) if not _is_empty(value)]
        if len(populated) < 2:
            continue
        text_count = sum(isinstance(row[column], str) for column in populated)
        following = _nonempty_rows_after(rows, index + 1, 30)
        if not following:
            continue
        reuse = sum(
            sum(not _is_empty(next_row[column]) for column in populated if column < len(next_row))
            / len(populated)
            for _, next_row in following
        ) / len(following)
        header_text = " ".join(str(row[column]) for column in populated)
        procedure_hints = sum(hint in header_text for hint in PROCEDURE_HEADER_HINTS)
        score = (
            len(populated)
            + text_count
            + reuse * len(populated)
            + procedure_hints * 5
            - index * 0.01
        )
        if score > best_score:
            best_index = index
            best_score = score
    if best_index is not None:
        return best_index
    return next(
        (index for index, row in enumerate(rows) if any(not _is_empty(value) for value in row)),
        None,
    )


def _active_columns(
        rows: Sequence[Sequence[Any]],
        header_index: int,
        sample_limit: int = 50,
) -> list[int]:
    maximum = max((len(row) for row in rows), default=0)
    sample = rows[header_index:header_index + sample_limit + 1]
    return [
        column
        for column in range(maximum)
        if any(column < len(row) and not _is_empty(row[column]) for row in sample)
    ]


def _sheet_density(
        rows: Sequence[Sequence[Any]],
        header_index: int,
        active_columns: Sequence[int],
) -> float:
    following = _nonempty_rows_after(rows, header_index + 1, 50)
    if not following or not active_columns:
        return 0.0
    ratios = [
        sum(column < len(row) and not _is_empty(row[column]) for column in active_columns)
        / len(active_columns)
        for _, row in following
    ]
    return float(median(ratios))


def _headers(
        header_row: Sequence[Any],
        active_columns: Sequence[int],
) -> dict[int, str]:
    result: dict[int, str] = {}
    seen: dict[str, int] = {}
    for column in active_columns:
        raw = header_row[column] if column < len(header_row) else None
        name = str(raw).strip() if not _is_empty(raw) else f"column_{get_column_letter(column + 1)}"
        key = name.casefold()
        if key in seen:
            name = f"{name}__{get_column_letter(column + 1)}"
        seen[key] = seen.get(key, 0) + 1
        result[column] = name
    return result


def _procedure_columns(
        rows: Sequence[Sequence[Any]],
        header_index: int,
        active_columns: Sequence[int],
) -> tuple[int, int] | None:
    """Find the numbering/title column pair used by procedure workbooks."""
    header_row = rows[header_index]
    labeled_columns = [
        column for column in active_columns
        if column < len(header_row) and not _is_empty(header_row[column])
    ]
    if not labeled_columns:
        return None

    title_column = labeled_columns[0]
    candidates: list[tuple[int, int]] = []
    for column in range(title_column):
        cooccurrence = sum(
            column < len(row)
            and title_column < len(row)
            and not _is_empty(row[column])
            and not _is_empty(row[title_column])
            for row in rows[header_index + 1:]
        )
        if cooccurrence >= 2:
            candidates.append((cooccurrence, column))
    if not candidates:
        return None
    _, index_column = max(candidates)
    return index_column, title_column


def _table_blocks(
        source_name: str,
        source_type: str,
        sheet_name: str,
        rows: Sequence[Sequence[Any]],
        header_index: int,
        active_columns: Sequence[int],
) -> list[dict[str, Any]]:
    headers = _headers(rows[header_index], active_columns)
    blocks: list[dict[str, Any]] = []
    for row_index in range(header_index + 1, len(rows)):
        row = rows[row_index]
        fields = {
            headers[column]: _normalize_value(row[column])
            for column in active_columns
            if column < len(row) and not _is_empty(row[column])
        }
        if not fields:
            continue
        blocks.append({
            "source_file": source_name,
            "source_type": source_type,
            "block_type": "table_row",
            "sheet": sheet_name,
            "row_start": row_index + 1,
            "row_end": row_index + 1,
            "fields": fields,
        })
    return blocks


def _procedure_blocks(
        source_name: str,
        source_type: str,
        sheet_name: str,
        rows: Sequence[Sequence[Any]],
        header_index: int,
        active_columns: Sequence[int],
) -> list[dict[str, Any]]:
    columns = _procedure_columns(rows, header_index, active_columns)
    if columns is None:
        return _table_blocks(
            source_name, source_type, sheet_name, rows, header_index, active_columns
        )
    index_column, title_column = columns
    header_row = rows[header_index]

    all_columns = sorted(set(active_columns) | {index_column, title_column})
    headers = _headers(header_row, all_columns)
    boundaries = [
        row_index
        for row_index in range(header_index + 1, len(rows))
        if index_column < len(rows[row_index])
        and title_column < len(rows[row_index])
        and not _is_empty(rows[row_index][index_column])
        and not _is_empty(rows[row_index][title_column])
    ]
    blocks: list[dict[str, Any]] = []
    section_title: str | None = None
    for boundary_index, row_start in enumerate(boundaries):
        row = rows[row_start]
        next_start = boundaries[boundary_index + 1] if boundary_index + 1 < len(boundaries) else len(rows)
        detail_columns = [
            column for column in all_columns
            if column not in {index_column, title_column}
        ]
        has_detail = any(
            column < len(row) and not _is_empty(row[column])
            for column in detail_columns
        )
        title = str(row[title_column]).strip()
        if not has_detail:
            section_title = title
            continue

        content_lines = []
        if section_title:
            content_lines.append(f"セクション: {section_title}")
        last_nonempty = row_start
        for row_index in range(row_start, next_start):
            current_row = rows[row_index]
            row_values = []
            for column in range(len(current_row)):
                if _is_empty(current_row[column]):
                    continue
                label = headers.get(column, f"column_{get_column_letter(column + 1)}")
                row_values.append(f"{label}: {current_row[column]}")
            if row_values:
                content_lines.append(" | ".join(row_values))
                last_nonempty = row_index

        path = [value for value in (section_title, title) if value]
        blocks.append({
            "source_file": source_name,
            "source_type": source_type,
            "block_type": "procedure_step",
            "sheet": sheet_name,
            "section_path": path,
            "row_start": row_start + 1,
            "row_end": last_nonempty + 1,
            "content": "\n".join(content_lines),
        })
    return blocks


def _sheet_blocks(
        source_name: str,
        source_type: str,
        sheet_name: str,
        rows: Sequence[Sequence[Any]],
        mode: str,
) -> list[dict[str, Any]]:
    header_index = _find_header_row(rows)
    if header_index is None:
        return []
    active_columns = _active_columns(rows, header_index)
    selected_mode = mode.casefold()
    if selected_mode == "auto":
        if _procedure_columns(rows, header_index, active_columns) is not None:
            selected_mode = "procedure"
        else:
            selected_mode = (
                "table" if _sheet_density(rows, header_index, active_columns) >= 0.60
                else "procedure"
            )
    if selected_mode == "procedure":
        return _procedure_blocks(
            source_name, source_type, sheet_name, rows, header_index, active_columns
        )
    if selected_mode != "table":
        raise ValueError(f"Unsupported Excel conversion mode: {mode}")
    return _table_blocks(
        source_name, source_type, sheet_name, rows, header_index, active_columns
    )


def excel_to_blocks(
        source_path: str | Path,
        excluded_sheets: str | Sequence[str] | None = DEFAULT_EXCLUDED_SHEETS,
        include_hidden: bool = False,
        mode: str = "auto",
        max_block_chars: int = DEFAULT_MAX_BLOCK_CHARS,
) -> list[dict[str, Any]]:
    source = Path(source_path)
    suffix = source.suffix.casefold()
    excluded = _parse_excluded_sheets(excluded_sheets)
    blocks: list[dict[str, Any]] = []

    if suffix == ".xlsx":
        formula_book = load_workbook(source, data_only=False, read_only=False, keep_links=False)
        value_book = load_workbook(source, data_only=True, read_only=False, keep_links=False)
        try:
            for formula_sheet in formula_book.worksheets:
                if formula_sheet.title.strip().casefold() in excluded:
                    continue
                if not include_hidden and formula_sheet.sheet_state != "visible":
                    continue
                value_sheet = value_book[formula_sheet.title]
                blocks.extend(_sheet_blocks(
                    source.name,
                    "excel",
                    formula_sheet.title,
                    _xlsx_rows(formula_sheet, value_sheet),
                    mode,
                ))
        finally:
            formula_book.close()
            value_book.close()
    elif suffix == ".xls":
        workbook = xlrd.open_workbook(source, on_demand=True)
        try:
            for sheet in workbook.sheets():
                if sheet.name.strip().casefold() in excluded:
                    continue
                if not include_hidden and getattr(sheet, "visibility", 0) != 0:
                    continue
                blocks.extend(_sheet_blocks(
                    source.name,
                    "excel",
                    sheet.name,
                    _xls_rows(sheet, workbook.datemode),
                    mode,
                ))
        finally:
            workbook.release_resources()
    elif suffix == ".csv":
        try:
            frame = pd.read_csv(source, header=None, encoding="utf-8")
        except UnicodeDecodeError:
            frame = pd.read_csv(source, header=None, encoding="cp932")
        blocks.extend(_sheet_blocks(
            source.name, "csv", source.stem, _dataframe_rows(frame), "table"
        ))
    else:
        raise ValueError("CSVまたはExcelファイルのみ対応しています")

    return split_evidence_blocks(blocks, max_block_chars)


def _iter_word_blocks(document) -> Iterator[tuple[int, Paragraph | Table]]:
    index = 0
    for child in document.element.body.iterchildren():
        if isinstance(child, CT_P):
            yield index, Paragraph(child, document)
            index += 1
        elif isinstance(child, CT_Tbl):
            yield index, Table(child, document)
            index += 1


def _heading_level(paragraph: Paragraph) -> int | None:
    style_name = paragraph.style.name if paragraph.style else ""
    match = re.search(r"(?:heading|見出し)\s*([1-9])", style_name, re.IGNORECASE)
    return int(match.group(1)) if match else None


def _word_table_blocks(
        source_name: str,
        section_path: Sequence[str],
        table: Table,
        table_index: int,
        body_index: int,
) -> list[dict[str, Any]]:
    rows = [[_normalize_value(cell.text.strip()) for cell in row.cells] for row in table.rows]
    if not rows:
        return []
    active_columns = list(range(max(len(row) for row in rows)))
    headers = _headers(rows[0], active_columns)
    blocks: list[dict[str, Any]] = []
    for row_index, row in enumerate(rows[1:], start=2):
        fields = {
            headers[column]: row[column]
            for column in active_columns
            if column < len(row) and not _is_empty(row[column])
        }
        if not fields:
            continue
        block = {
            "source_file": source_name,
            "source_type": "word",
            "block_type": "word_table_row",
            "section_path": list(section_path),
            "table_index": table_index,
            "row_start": row_index,
            "row_end": row_index,
            "body_index": body_index,
            "fields": fields,
        }
        if not block["section_path"]:
            block.pop("section_path")
        blocks.append(block)
    return blocks


def word_to_blocks(
        source_path: str | Path,
        max_block_chars: int = DEFAULT_MAX_BLOCK_CHARS,
) -> list[dict[str, Any]]:
    source = Path(source_path)
    if source.suffix.casefold() != ".docx":
        raise ValueError("DOCXファイルのみ対応しています")

    document = Document(source)
    heading_path: list[str] = []
    section_units: list[tuple[int, str]] = []
    blocks: list[dict[str, Any]] = []
    table_index = 0

    def flush_section():
        if not section_units:
            return
        block = {
            "source_file": source.name,
            "source_type": "word",
            "block_type": "word_section",
            "section_path": list(heading_path),
            "row_start": section_units[0][0] + 1,
            "row_end": section_units[-1][0] + 1,
            "content": "\n".join(text for _, text in section_units),
        }
        if not block["section_path"]:
            block.pop("section_path")
        blocks.append(block)
        section_units.clear()

    for body_index, item in _iter_word_blocks(document):
        if isinstance(item, Paragraph):
            style_name = item.style.name if item.style else ""
            if style_name.casefold().startswith("toc"):
                continue
            text = item.text.strip()
            if not text:
                continue
            level = _heading_level(item)
            if level is not None:
                flush_section()
                heading_path[:] = heading_path[:level - 1] + [text]
                continue
            is_numbered = item._p.pPr is not None and item._p.pPr.numPr is not None
            is_list_style = (
                "list" in style_name.casefold()
                or "箇条書き" in style_name
                or "番号" in style_name
            )
            if is_numbered or is_list_style:
                text = f"- {text}"
            section_units.append((body_index, text))
        else:
            flush_section()
            table_index += 1
            blocks.extend(_word_table_blocks(
                source.name, heading_path, item, table_index, body_index
            ))
    flush_section()
    return split_evidence_blocks(blocks, max_block_chars)
